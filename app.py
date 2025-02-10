# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.
import numpy as np
import pandas as pd
from flask import Flask, request, jsonify, render_template
import pickle
import joblib
from openpyxl import load_workbook, Workbook
flask_app = Flask(__name__)

# with open('rf_model (1).pkl', 'rb') as f:
#     model_1 = pickle.load(f)
model_1 = joblib.load('rf_model.pkl')

with open('xgb_model.pkl', 'rb') as f:
    model_2 = pickle.load(f)

# model_3 = joblib.load('stacked_model.pkl')

models = {
    "model_1": model_1,
    "model_2": model_2,
    # "model_3": model_3
}

EXCEL_FILE = "loan_predictions.xlsx"

def extract_year(date_string):
    return int(date_string.split('-')[0])  # Extracts the year (e.g., 2005 from "2005-06")

@flask_app.route("/")
def Home():
    return render_template("index.html")

dummies = ['verification_status', 'purpose',
           'application_type', 'home_ownership', 'sub_grade']

@flask_app.route('/predict', methods=['POST'])
def predict():
    try:
        # Get form data
        form_data = request.form

        selected_model = form_data['model']
        model = models[selected_model]
        # print(type(model))

        input_dict = {
            'loan_amnt': float(form_data['loan_amnt']),
            'term': float(form_data['term']),
            'int_rate': float(form_data['int_rate']),
            'installment': float(form_data['installment']),
            'annual_inc': float(form_data['annual_inc']),
            'dti': float(form_data['dti']),
            'earliest_cr_line': extract_year(form_data['earliest_cr_line']),
            'revol_bal': float(form_data['revol_bal']),
            'revol_util': float(form_data['revol_util']),
            'mort_acc': float(form_data['mort_acc']),
            'avg_cur_bal': float(form_data['avg_cur_bal']),
            'delinq_2yrs': float(form_data['delinq_2yrs']),
            'verification_status': form_data['verification_status'],
            'purpose': form_data['purpose'],
            'application_type': form_data['application_type'],
            'home_ownership': form_data['home_ownership'],
            'sub_grade': form_data['sub_grade']
        }

        # Convert input_dict to a DataFrame
        input_df = pd.DataFrame([input_dict])

        # Create dummies for categorical columns
        input_df = pd.get_dummies(input_df, columns=dummies, drop_first=True)

        # Align with the training feature set
        with open('final_features.pkl', 'rb') as f:
            final_features = pickle.load(f)
        # print(final_features)

        for col in final_features:
            if col not in input_df.columns:
                input_df[col] = 0

        input_df = input_df[final_features]
        # print("Input features:", input_df.columns.tolist())

        # prediction = model.predict(input_df.values)
        #
        # result = "Default Risk" if prediction[0] == 1 else "No Default Risk"
        # return render_template('index.html', prediction_text=result)
        probabilities = model.predict_proba(input_df.values)

        prob_no_default = probabilities[0][0] * 100
        prob_default = probabilities[0][1] * 100
        save_to_excel(input_dict, prob_no_default, prob_default)

        return jsonify(
            prob_no_default = round(prob_no_default,2),
            prob_default = round(prob_default,2)
        )
    except Exception as e:
        print("Error:", e)
        return jsonify( error_message="Error occurred during prediction.")

def save_to_excel(data, prob_no_default, prob_default):
    """Append form data and predictions to an Excel file without checking for file existence."""
    try:
        wb = load_workbook(EXCEL_FILE)  # Try to load the existing Excel file
        ws = wb.active
    except:
        wb = Workbook()  # Create a new workbook if the file doesn't exist
        ws = wb.active
        # Write the header row
        ws.append(list(data.keys()) + ["prob_no_default", "prob_default"])

    # Append new data row
    ws.append(list(data.values()) + [round(prob_no_default, 2), round(prob_default, 2)])

    # Save the file
    wb.save(EXCEL_FILE)

if __name__ == "__main__":
    flask_app.run(debug=True)